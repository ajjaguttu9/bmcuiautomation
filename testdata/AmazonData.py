import openpyxl,os
# import deepcopy module
from copy import deepcopy

class AmazonData:

    @staticmethod
    def getTestData():
        alldata = []
        cell_data = {}
        cwd = os.getcwd()
        book = openpyxl.load_workbook(cwd+"\\testdata\\AmazonUI_Test.xlsx")
        sheet = book.active
        for row in range(2, sheet.max_row + 1):
            for col in range(2, sheet.max_column + 1):  # to get columns
                cell_data[sheet.cell(row=1, column=col).value] = sheet.cell(row=row, column=col).value
            alldata.append(deepcopy(cell_data))
        return alldata
